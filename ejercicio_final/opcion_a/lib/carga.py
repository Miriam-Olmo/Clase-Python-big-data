import csv
from openpyxl import load_workbook
import os
import json
import xml.etree.ElementTree as et


def mostrar_info(nombre_fichero, columnas, registros):

    print("\n" + "=" * 50)
    print("FICHERO:", nombre_fichero)
    print("=" * 50)

    # Número total de registros
    print("\nNúmero total de registros:", len(registros))

    # Campos / columnas
    print("\nCampos/Columnas:")
    print(columnas)

    # Primeros 5 registros
    print("\nPrimeros 5 registros:")

    for registro in registros[:5]:
        print(registro)




def leer_csv(carpeta, archivo):
    fichero = open(f"{carpeta}/{archivo}", "r", encoding='UTF-8')
    lector = csv.DictReader(fichero) 
    artistas = list(lector)
    fichero.close()
    return artistas
leer_csv('datos', 'artistas.csv')




def cargar_excel(carpeta, archivo):
   excel = load_workbook(f"./{carpeta}/{archivo}")
   hoja = excel.active
   
   filas = hoja.iter_rows(values_only = True)
   cabeceras = next(filas)
   
   lista_resultante = []
   for fila in hoja.iter_rows(min_row=2, values_only=True):
       escenario = dict(zip(cabeceras, fila))
       lista_resultante.append(escenario)
       
   return lista_resultante

cargar_excel('datos', 'escenarios_horarios.xlsx')




def cargar_json(carpeta, nombre):
    fichero = open(f"./{carpeta}/{nombre}", "r", encoding="UTF-8")
    datos = json.load(fichero)
    fichero.close()
    return datos

cargar_json('datos', 'ventas_entradas.json')





def cargar_xml(carpeta, fichero):
    # leer un archivo en bruto
    fichero = et.parse(f'./{carpeta}/{fichero}')
    # me devuelve el nodo principal
    nodo_raiz = fichero.getroot()

    patrocinadores = []

    for patrocinador in nodo_raiz.findall('patrocinador'):
        patrocinador = {
            'nombre_empresa': patrocinador.find('nombre_empresa').text,
            'contacto': patrocinador.find('contacto').text,
            'email': patrocinador.find('email').text,
            'importe_patrocinio': patrocinador.find('importe_patrocinio').text,
            'categoria': patrocinador.find('categoria').text,
            'fecha_inicio': patrocinador.find('fecha_inicio').text,
            'fecha_fin': patrocinador.find('fecha_fin').text
        }
        patrocinadores.append(patrocinador)
    return patrocinadores
    
cargar_xml('datos', 'patrocinadores.xml')


def resumen(nombre, datos):

    print("\n==========")
    print(nombre)
    print("==========")

    print("Total registros:", len(datos))

    print("\nCampos:")

    for campo in datos[0].keys():
        print("-", campo)

    print("\nPrimeros 5 registros:")

    for registro in datos[:5]:
        print(registro)

artistas = leer_csv('datos', 'artistas.csv')
programa = cargar_excel('datos', 'escenarios_horarios.xlsx')
ventas = cargar_json('datos', 'ventas_entradas.json')
patrocinadores = cargar_xml('datos', 'patrocinadores.xml')
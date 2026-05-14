from openpyxl import load_workbook


# cargar el fichero de excel en nuestro archivo
excel = load_workbook('./data/empleados.xlsx')
# cuando tengo el excel tengo que seleccionar la hoja, en este caso la hoja activa
hoja = excel.active

# recorrer las filas y columnas de la hoja
for fila in hoja.iter_rows(min_row=2, values_only=True):
    print(fila)